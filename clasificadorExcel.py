print("}The program is loading, please wait...")

from tkinter import Button, Entry, Frame, Label, Tk
from tkinter import ttk
from tkinter.constants import BOTH, LEFT, X, Y
import openpyxl
from openpyxl import workbook
from openpyxl.cell import cell
from openpyxl.workbook.workbook import Workbook
from tkinter import filedialog


path= filedialog.askopenfilename(initialdir="/", title="Please select the file to filter", filetypes=(("Excel Files", ".xlsx"),("Todos los archivos", ".*")))
OriginExcel = openpyxl.load_workbook(path)
OriginPage = OriginExcel.active
excelResultados = Workbook() 
paginaResultados=excelResultados.active 
#Variables de openpyxl, define la ubicación del archivo y la OriginPage activa
tituloTXT="Filtro de la lista de Atena\nBienvenido al filtro de la lista de Atena, favor de poner sus parametros a filtrar"
indicacion1TXT="Palabra clave a buscar:"
row = 0
opcionesDesplegable=[]
antiloop=0
find=True
while find:
    row +=1
    for i in range (1, OriginPage.max_column+1):
        nombreColumna=OriginPage.cell(row=row, column=i)
        if nombreColumna.value is None:
            print("Vacio")
            if antiloop>=20:
                print("Antiloop activado")
                break
            continue
        else:
            opcionesDesplegable.append(nombreColumna.value)
            find = False
especificacionesDesplegable="Selecciona en que criterio quieres buscar\n En caso de no tener un criterio especifico, dejar vacía la caja"
graciasTXT="Gracias por usar nuestro programa\n Se generó un archivo de excel con los resultados en la carpeta\n Para buscar otra cosa favor de usar este boton"
#Aqui modificamos los textos de nuestras etiquetas, botones etc
busquedaConcluyente=[1]
#Variables normales para el código

def seleccionarBusqueda():
    seleccion=desplegable.current() #la variable es el valor del indice
    if seleccion==-1: #En caso de que no hayamos seleccionado nada, el valor del indice es -1
        busquedaSinEspecificar()
    else:
        busquedaEspecificada(seleccion)

def busquedaEspecificada(select):
    global row
    global busquedaConcluyente
    busqueda = entryFiltro.get()
    print(busqueda)
    for i in range(2, OriginPage.max_row+1):
        celda = OriginPage.cell(row = row + i , column = (select+1))
        if celda.value == busqueda.strip():
            busquedaConcluyente.append(i)
            print(celda.value)
    print("Numero de resultados: {}".format(len(busquedaConcluyente)))
    imprimir()


def busquedaSinEspecificar():
    global row
    contador = 1
    busqueda = entryFiltro.get()
    global busquedaConcluyente
    while contador <= 5:
        for i in range(2, OriginPage.max_row+1):
            celda = OriginPage.cell(row = row + i , column = contador)
            if celda.value == busqueda.strip():
                busquedaConcluyente.append(i)
        contador += 1
    print("Numero de resultados: {}".format(len(busquedaConcluyente)))
    imprimir()

def imprimir():
    global busquedaConcluyente
    contadorFila=1
    contadorColumna=1
    for i in busquedaConcluyente:
        fila = OriginPage[i]
        for n in fila:
            n.value
            NuevaCelda=paginaResultados.cell(row=contadorFila, column=contadorColumna)
            NuevaCelda.value=n.value 
            contadorColumna+=1
        contadorColumna=1    
        contadorFila+=1
        print("{}/{}".format(contadorFila,len(busquedaConcluyente)))
    camino="{}/resultados.xlsx".format(filedialog.askdirectory(initialdir="/", title="Seleccione donde desea el archivo de resultados"))
    print(camino)
    excelResultados.save(camino)
    filtro.pack_forget()
    resultados.pack(fill=BOTH , expand=True)
    gracias.place(x=45, y=30)
    boton2.place(x=232, y=150)

def limpiar():
    resultados.pack_forget()
    filtro.pack(fill=BOTH , expand=True)
    global busquedaConcluyente
    busquedaConcluyente.clear()
    busquedaConcluyente.append(1)
    

root = Tk()
root.title("filtro lista de atena")
root.geometry("750x750")
root.config(bg="#84DC76")
#Caracteristicas generales de la ventana

filtro = Frame(root, bg="#84DC76")
resultados = Frame(root, bg="#84DC76")
#Aquí declaramos las pestañas, junto con sus parametros

titulo=Label(filtro, text=tituloTXT, bg="#84DC76", fg="#FFFFFF", font="HP 16")
boton = Button(filtro, text = "Buscar", height=1, width=10, font="HP 20", command= seleccionarBusqueda)
entryFiltro = Entry(filtro, font="HP 16")
titulo=Label(filtro, text=tituloTXT, bg="#84DC76", fg="#FFFFFF", font="HP 16")
especificacionesDesplegable=Label(filtro, text=especificacionesDesplegable, bg="#84DC76", fg="#FFFFFF", font="HP 16")
indicacion1=Label(filtro, text=indicacion1TXT, bg="#84DC76", fg="#FFFFFF", font="HP 14")
desplegable =ttk.Combobox(filtro, width=37, height=60, state="readonly")
gracias=Label(resultados, text=graciasTXT, bg="#84DC76", fg="#FFFFFF", font="HP 16")
boton2 = Button(resultados, text = "Buscar de nuevo", height=2, width=15, font="HP 16", command= limpiar)

#Aquí declaramos las etiquetas, botones, entradas de texto, etc

desplegable["values"]=opcionesDesplegable
filtro.pack(fill=BOTH , expand=True)
titulo.place(x=7, y=30)
indicacion1.place(x=255, y=100)
entryFiltro.place(x=235, y=130)
boton.place(x=275, y=300)
especificacionesDesplegable.place(x=75, y=185)
desplegable.place(x=235, y=250)
#Aqui llamamos a las variables de las etiquetas para que aparezcan


root.mainloop()

